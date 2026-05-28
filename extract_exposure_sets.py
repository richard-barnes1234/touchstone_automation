import pandas as pd
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
 
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
  
from api_client import send_soap_request
from parse_response import parse_xml
from config import BUSINESS_UNIT_SID, SQL_INSTANCE_SID, DATA_SOURCE_SID


soap_body = """<s:Envelope xmlns:s="http://www.w3.org/2003/05/soap-envelope" xmlns:a="http://www.w3.org/2005/08/addressing">
        <s:Header>
            <a:Action s:mustUnderstand="1">AIR.Services.DataSourceManagementService.Api/IDataSourceManagementService/GetExposureSets</a:Action>
            <a:MessageID>urn:uuid:f23aa50c-9a41-4932-bd0e-51a37913fde1</a:MessageID>
            <a:ReplyTo>
                    <a:Address>http://www.w3.org/2005/08/addressing/anonymous</a:Address>
            </a:ReplyTo>
            <a:To s:mustUnderstand="1">https://crcins-na-prod-touchstoneapi.air-worldwide.com/FEP/AirServiceFacade.svc</a:To>
        </s:Header>
        <s:Body>
            <GetExposureSets xmlns="AIR.Services.DataSourceManagementService.Api">
            <request xmlns:b="AIR.Services.DataSourceManagement.Api" xmlns:i="http://www.w3.org/2001/XMLSchema-instance">
            <BusinessUnitSid xmlns="AIR.Services.Common.Api">1</BusinessUnitSid>
            <LicenseUid xmlns="AIR.Services.Common.Api">00000000-0000-0000-0000-000000000000</LicenseUid>
            <RequestUid xmlns="AIR.Services.Common.Api">00000000-0000-0000-0000-000000000000</RequestUid>
            <SqlInstanceSid xmlns="AIR.Services.Common.Api">1</SqlInstanceSid>
            <b:DataSourceSid>2</b:DataSourceSid>
            </request>
            </GetExposureSets>
        </s:Body>  
</s:Envelope>""" 

def format_worksheet(ws):
        """Apply professional formatting to the worksheet in-place."""
            
        HEADER_BG   = "1F3864"
        HEADER_FONT = "FFFFFF"
        ROW_ALT_BG  = "EAF0FB"
        BORDER_CLR  = "B8CCE4"          
        
        thin = Border(
                left   = Side(style="thin", color=BORDER_CLR),
                right  = Side(style="thin", color=BORDER_CLR),
                top    = Side(style="thin", color=BORDER_CLR),
                bottom = Side(style="thin", color=BORDER_CLR),
        )

        for cell in ws[1]:
            cell.font      = Font(name="Arial", bold=True, color=HEADER_FONT, size=10)
            cell.fill      = PatternFill("solid", start_color=HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = thin
        ws.row_dimensions[1].height = 30

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                bg = ROW_ALT_BG if row_idx % 2 == 0 else "FFFFFF"
                for cell in row:
                    cell.font      = Font(name="Arial", size=9)
                    cell.fill      = PatternFill("solid", start_color=bg)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border    = thin

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)
            
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_properties.tabColor = "1F3864"

print("Fetching Exposure Sets...")

def extract_exposure_sets_data():
    """Fetches exposure sets from Touchstone and returns a DataFrame"""
    print("Fetching Exposure Sets...")
    response = send_soap_request(soap_body)
    print(f"Status Code: {response.status_code}")

    if response.status_code == 200:
        print("Success!")
        from parse_exposure_sets import parse_exposure_sets, print_exposure_sets
        exposure_sets = parse_exposure_sets(response.text)
        print_exposure_sets(exposure_sets)

        df = pd.DataFrame(exposure_sets)

        # Save to Excel
        OUTPUT = "exposure_sets.xlsx"
        df.to_excel(OUTPUT, index=False, sheet_name="Exposure Sets")
        wb = load_workbook(OUTPUT)
        format_worksheet(wb["Exposure Sets"])
        wb.save(OUTPUT)
        print(f"Saved {len(exposure_sets)} exposure sets to {OUTPUT}")

        return df
    else:
        print("Failed!")
        print(response.text)
        return None

if __name__ == "__main__":
    extract_exposure_sets_data()