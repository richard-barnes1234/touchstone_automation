# app.py — Touchstone Results Dashboard

import streamlit as st
import pandas as pd
import urllib3
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

st.set_page_config(
    page_title="Touchstone Results",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 2rem 2.5rem; max-width: 1300px; }
section[data-testid="stSidebar"] { background: #0D1B2A; border-right: 1px solid #1E3448; }
section[data-testid="stSidebar"] * { color: #A8BFCF !important; }
section[data-testid="stSidebar"] .stRadio label { font-size: 14px !important; padding: 8px 12px !important; border-radius: 6px !important; cursor: pointer !important; }
section[data-testid="stSidebar"] .stRadio label:hover { background: #1E3448 !important; color: #E8F4FD !important; }
.page-header { padding: 1.5rem 0 1rem 0; border-bottom: 1px solid #E2EAF4; margin-bottom: 1.5rem; }
.page-header h1 { font-size: 1.6rem; font-weight: 600; color: #0D1B2A; margin: 0; }
.page-header p  { font-size: 0.875rem; color: #6B7E8F; margin: 0.25rem 0 0 0; }
.card { background: #fff; border: 1px solid #E2EAF4; border-radius: 10px; padding: 1.25rem 1.5rem; margin-bottom: 1rem; box-shadow: 0 1px 3px rgba(13,27,42,0.04); }
.card-title { font-size: 0.75rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.08em; color: #6B7E8F; margin-bottom: 0.5rem; }
.card-value { font-size: 1.75rem; font-weight: 600; color: #0D1B2A; line-height: 1; }
.card-sub   { font-size: 0.8rem; color: #6B7E8F; margin-top: 0.35rem; }
.metric-row { display: grid; grid-template-columns: repeat(3, 1fr); gap: 1rem; margin-bottom: 1.5rem; }
.metric-card { background: #fff; border: 1px solid #E2EAF4; border-radius: 10px; padding: 1.2rem 1.4rem; box-shadow: 0 1px 3px rgba(13,27,42,0.04); }
.metric-card.blue  { border-left: 3px solid #3B82F6; }
.metric-card.green { border-left: 3px solid #22C55E; }
.metric-card.amber { border-left: 3px solid #F59E0B; }
.metric-card .label { font-size: 0.7rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.1em; color: #6B7E8F; margin-bottom: 0.4rem; }
.metric-card .value { font-size: 1.5rem; font-weight: 600; color: #0D1B2A; }
.metric-card .sub   { font-size: 0.75rem; color: #6B7E8F; margin-top: 0.2rem; }
.project-row { display: flex; align-items: center; justify-content: space-between; padding: 0.65rem 1rem; border: 1px solid #E2EAF4; border-radius: 8px; margin-bottom: 0.4rem; background: #fff; cursor: pointer; transition: all 0.15s; }
.project-row:hover { background: #EFF6FF; border-color: #BFDBFE; }
.project-row.selected { background: #EFF6FF; border-color: #3B82F6; border-left: 3px solid #3B82F6; }
.project-name { font-size: 0.875rem; font-weight: 500; color: #0D1B2A; }
.project-sid  { font-family: 'DM Mono', monospace; font-size: 0.75rem; color: #6B7E8F; }
.project-meta { font-size: 0.75rem; color: #9CA3AF; }
.section-header { font-size: 0.85rem; font-weight: 600; color: #374151; margin: 1.25rem 0 0.75rem 0; padding-bottom: 0.5rem; border-bottom: 1px solid #F1F5F9; }
.stButton > button { background: #1E40AF !important; color: #fff !important; border: none !important; border-radius: 7px !important; padding: 0.5rem 1.25rem !important; font-size: 0.875rem !important; font-weight: 500 !important; }
.stButton > button:hover { background: #1D3FA0 !important; }
.page-btn > button { background: #F1F5F9 !important; color: #374151 !important; border: 1px solid #E2EAF4 !important; font-size: 0.8rem !important; padding: 0.35rem 0.75rem !important; }
</style>
""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="padding:1.5rem 1rem 1rem 1rem;">
        <div style="font-size:1.1rem;font-weight:600;color:#E8F4FD;">Touchstone</div>
        <div style="font-size:0.75rem;color:#4A6580;margin-top:2px;">Results Dashboard</div>
    </div>
    """, unsafe_allow_html=True)
    page = st.radio("", ["🏠  Home", "📁  Projects", "📊  Results", "📋  History"], label_visibility="collapsed")
    st.markdown("""
    <div style="position:absolute;bottom:1.5rem;left:1rem;right:1rem;">
        <div style="font-size:0.7rem;color:#2D4559;border-top:1px solid #1E3448;padding-top:0.75rem;">
            Touchstone AIR Cloud · SOAP API
        </div>
    </div>
    """, unsafe_allow_html=True)

PAGE_SIZE = 50

# ── Excel builder ─────────────────────────────────────────────────────────────
def format_sheet(ws):
    HEADER_BG = "1F3864"; HEADER_FG = "FFFFFF"; ALT_BG = "EAF0FB"; BORDER_CLR = "B8CCE4"
    thin = Border(
        left=Side(style="thin", color=BORDER_CLR), right=Side(style="thin", color=BORDER_CLR),
        top=Side(style="thin", color=BORDER_CLR),  bottom=Side(style="thin", color=BORDER_CLR)
    )
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color=HEADER_FG, size=10)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 30
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        bg = ALT_BG if row_idx % 2 == 0 else "FFFFFF"
        for cell in row:
            cell.font = Font(name="Arial", size=9)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_excel(analysis_sid, analysis_name, project_name, datasets):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Touchstone Loss Analysis Report"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color="1F3864")
    ws["A3"] = "Project";       ws["B3"] = project_name
    ws["A4"] = "Analysis SID";  ws["B4"] = analysis_sid
    ws["A5"] = "Analysis Name"; ws["B5"] = analysis_name
    ws["A6"] = "Generated";     ws["B6"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A8"] = "Contents"
    row = 9
    for name, df in datasets.items():
        if df is not None and not df.empty:
            ws[f"A{row}"] = f"  • {name}"; ws[f"B{row}"] = f"{len(df):,} records"; row += 1
    for col in ["A", "B"]: ws.column_dimensions[col].width = 30
    for sheet_name, df in datasets.items():
        if df is not None and not df.empty:
            ws_data = wb.create_sheet(title=sheet_name[:31])
            for col_idx, col_name in enumerate(df.columns, start=1):
                ws_data.cell(row=1, column=col_idx, value=col_name)
            for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws_data.cell(row=row_idx, column=col_idx, value=value)
            format_sheet(ws_data)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def log_pull(project_name, analysis_sid, analysis_name, row_counts):
    if "history" not in st.session_state:
        st.session_state.history = []
    st.session_state.history.insert(0, {
        "Timestamp":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Project":       project_name,
        "Analysis SID":  analysis_sid,
        "Analysis Name": analysis_name,
        "ELT Rows":      row_counts.get("ELT", 0),
        "EP Rows":       row_counts.get("EP Curves", 0),
        "Summary Rows":  row_counts.get("Loss Summary", 0),
    })


# ════════════════════════════════════════════════════════════════════════════
# HOME
# ════════════════════════════════════════════════════════════════════════════
if page == "🏠  Home":
    st.markdown("""
    <div class="page-header">
        <h1>Touchstone Results Dashboard</h1>
        <p>Pull catastrophe model results directly from Touchstone — no manual copy-paste required</p>
    </div>""", unsafe_allow_html=True)

    st.markdown('<div class="section-header">How to use</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="card">
        <ol style="margin:0;padding-left:1.2rem;font-size:0.875rem;color:#374151;line-height:2.2;">
            <li>Go to <strong>Projects</strong> — load and search across 2,252 projects</li>
            <li>Select a project — see its completed analyses (first 50 shown)</li>
            <li>Select an analysis — go to <strong>Results</strong></li>
            <li>Click <strong>Fetch Results</strong> — ELT, EP Curves, Loss Summary pulled automatically</li>
            <li>Click <strong>Download Excel Report</strong></li>
        </ol>
    </div>""", unsafe_allow_html=True)

    projects = st.session_state.get("all_projects", [])
    history  = st.session_state.get("history", [])
    st.markdown(f"""
    <div class="metric-row">
        <div class="metric-card blue">
            <div class="label">Projects loaded</div>
            <div class="value">{len(projects):,}</div>
            <div class="sub">From Touchstone API</div>
        </div>
        <div class="metric-card green">
            <div class="label">Reports pulled</div>
            <div class="value">{len(history)}</div>
            <div class="sub">This session</div>
        </div>
        <div class="metric-card amber">
            <div class="label">Data source</div>
            <div class="value">SOAP API</div>
            <div class="sub">Touchstone AIR Cloud</div>
        </div>
    </div>""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════════
# PROJECTS
# ════════════════════════════════════════════════════════════════════════════
elif page == "📁  Projects":
    st.markdown("""
    <div class="page-header">
        <h1>Projects</h1>
        <p>Browse and search all projects — select one to view its analyses</p>
    </div>""", unsafe_allow_html=True)

    # ── Load projects ─────────────────────────────────────────────────────────
    col1, col2 = st.columns([1, 4])
    with col1:
        if st.button("🔄  Load Projects"):
            with st.spinner("Fetching all projects from Touchstone..."):
                try:
                    from get_analysis_sids import get_all_projects
                    projects = get_all_projects()
                    st.session_state["all_projects"]  = projects
                    st.session_state["project_page"]  = 0
                    st.session_state["project_search"] = ""
                    st.success(f"✓ {len(projects):,} projects loaded")
                except Exception as e:
                    st.error(f"Failed: {e}")

    projects = st.session_state.get("all_projects", [])

    if projects:
        # ── Search ────────────────────────────────────────────────────────────
        st.markdown('<div class="section-header">Search Projects</div>', unsafe_allow_html=True)
        search = st.text_input(
            "Search", placeholder="Type project name...",
            label_visibility="collapsed",
            key="project_search_input"
        )

        # Filter by name
        if search:
            filtered = [p for p in projects if search.lower() in p["ProjectName"].lower()]
            if st.session_state.get("last_search") != search:
                st.session_state["project_page"] = 0
                st.session_state["last_search"]  = search
        else:
            filtered = projects

        # ── Pagination ────────────────────────────────────────────────────────
        total_pages = max(1, (len(filtered) + PAGE_SIZE - 1) // PAGE_SIZE)
        current_page = st.session_state.get("project_page", 0)
        current_page = min(current_page, total_pages - 1)

        start = current_page * PAGE_SIZE
        end   = start + PAGE_SIZE
        page_projects = filtered[start:end]

        # Stats row
        st.markdown(f"""
        <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:0.75rem;">
            <span style="font-size:0.8rem;color:#6B7E8F;">
                Showing <strong>{start+1}–{min(end, len(filtered))}</strong> of <strong>{len(filtered):,}</strong> projects
                {"(filtered)" if search else ""}
            </span>
            <span style="font-size:0.8rem;color:#6B7E8F;">Page {current_page+1} of {total_pages}</span>
        </div>""", unsafe_allow_html=True)

        # ── Project list ──────────────────────────────────────────────────────
        selected_project_sid  = st.session_state.get("selected_project_sid")
        selected_project_name = st.session_state.get("selected_project_name", "")

        for p in page_projects:
            is_selected = str(p["ProjectSid"]) == str(selected_project_sid)
            card_class  = "project-row selected" if is_selected else "project-row"
            result_text = f"{p['ResultCount']} result(s)" if p.get('ResultCount') else ""

            col_name, col_btn = st.columns([5, 1])
            with col_name:
                st.markdown(f"""
                <div class="{card_class}">
                    <div>
                        <div class="project-name">{p["ProjectName"]}</div>
                        <div class="project-sid">SID: {p["ProjectSid"]} &nbsp;·&nbsp; {result_text}</div>
                    </div>
                </div>""", unsafe_allow_html=True)
            with col_btn:
                if st.button("Select", key=f"sel_{p['ProjectSid']}"):
                    st.session_state["selected_project_sid"]       = p["ProjectSid"]
                    st.session_state["selected_project_name"]      = p["ProjectName"]
                    st.session_state["project_analyses"]           = None
                    st.session_state["selected_analysis_sid"]      = None
                    st.session_state["selected_analysis_name"]     = None
                    st.rerun()

        # ── Pagination controls ───────────────────────────────────────────────
        st.markdown("<div style='margin-top:1rem;'></div>", unsafe_allow_html=True)
        col_prev, col_info, col_next = st.columns([1, 2, 1])
        with col_prev:
            if current_page > 0:
                if st.button("← Previous 50"):
                    st.session_state["project_page"] = current_page - 1
                    st.rerun()
        with col_info:
            st.markdown(f"<div style='text-align:center;font-size:0.8rem;color:#6B7E8F;padding-top:0.5rem;'>Page {current_page+1} of {total_pages}</div>", unsafe_allow_html=True)
        with col_next:
            if current_page < total_pages - 1:
                if st.button("Next 50 →"):
                    st.session_state["project_page"] = current_page + 1
                    st.rerun()

        # ── Selected project — show analyses ──────────────────────────────────
        if selected_project_sid:
            st.markdown(f'<div class="section-header">Analyses — {selected_project_name}</div>', unsafe_allow_html=True)

            col1, col2 = st.columns([1, 4])
            with col1:
                if st.button("📋  Load Analyses"):
                    with st.spinner(f"Fetching analyses for {selected_project_name}..."):
                        try:
                            from get_analysis_sids import get_analyses_for_project, get_hazard_analyses_for_project
                            loss_analyses = get_analyses_for_project(selected_project_sid, selected_project_name)
                            haz_analyses  = get_hazard_analyses_for_project(selected_project_sid, selected_project_name)
                            all_analyses  = loss_analyses + haz_analyses
                            st.session_state["project_analyses"] = all_analyses
                            st.success(f"✓ {len(loss_analyses)} loss + {len(haz_analyses)} hazard = {len(all_analyses)} total analyses")
                        except Exception as e:
                            st.error(f"Failed: {e}")

            analyses = st.session_state.get("project_analyses")

            if analyses is not None:
                if len(analyses) == 0:
                    st.info("No completed analyses found for this project.")
                else:
                    # Search box
                    analysis_search = st.text_input(
                        "Search analyses",
                        placeholder="Search by Analysis SID or name...",
                        label_visibility="collapsed",
                        key="analysis_search_input"
                    )

                    # Filter by SID or name
                    if analysis_search:
                        filtered_analyses = [
                            a for a in analyses
                            if analysis_search.lower() in a["AnalysisName"].lower()
                            or analysis_search in str(a["AnalysisSid"])
                        ]
                    else:
                        filtered_analyses = analyses

                    # Show first 50 of filtered
                    shown = filtered_analyses[:50]
                    st.markdown(f"<div style='font-size:0.8rem;color:#6B7E8F;margin-bottom:0.75rem;'>Showing {len(shown)} of {len(filtered_analyses)} analyses{' (filtered)' if analysis_search else ''}</div>", unsafe_allow_html=True)

                    for a in shown:
                        col_a, col_b = st.columns([5, 1])
                        with col_a:
                            a_type      = a.get("AnalysisType", "LOSS")
                            badge_color = "#DBEAFE" if a_type == "LOSS" else "#FEF3C7"
                            badge_text  = "#1E40AF" if a_type == "LOSS" else "#92400E"
                            st.markdown(f"""
                            <div class="project-row">
                                <div style="display:flex;align-items:center;gap:0.75rem;">
                                    <span style="background:{badge_color};color:{badge_text};font-size:0.7rem;font-weight:600;padding:2px 8px;border-radius:4px;">{a_type}</span>
                                    <div>
                                        <div class="project-name">{a["AnalysisName"]}</div>
                                        <div class="project-sid">SID: {a["AnalysisSid"]} &nbsp;·&nbsp; Completed: {a.get("Completed","—")}</div>
                                    </div>
                                </div>
                            </div>""", unsafe_allow_html=True)
                        with col_b:
                            if st.button("Select", key=f"ana_{a['AnalysisSid']}"):
                                st.session_state["selected_analysis_sid"]  = a["AnalysisSid"]
                                st.session_state["selected_analysis_name"] = a["AnalysisName"]
                                st.session_state["last_results"]           = None
                                st.rerun()

                    if st.session_state.get("selected_analysis_sid"):
                        st.success(f"✓ Analysis selected — SID {st.session_state['selected_analysis_sid']} — go to **Results** to fetch data")

    else:
        st.info("Click **Load Projects** to fetch all projects from Touchstone")


# ════════════════════════════════════════════════════════════════════════════
# RESULTS
# ════════════════════════════════════════════════════════════════════════════
elif page == "📊  Results":
    st.markdown("""
    <div class="page-header">
        <h1>Results</h1>
        <p>Fetch and download loss data for the selected analysis</p>
    </div>""", unsafe_allow_html=True)

    analysis_sid  = st.session_state.get("selected_analysis_sid")
    analysis_name = st.session_state.get("selected_analysis_name", "")
    project_name  = st.session_state.get("selected_project_name", "")

    if not analysis_sid:
        st.info("No analysis selected — go to **Projects** to select one first.")
    else:
        st.markdown(f"""
        <div class="card">
            <div class="card-title">Selected Analysis</div>
            <div class="card-value" style="font-size:1.2rem;">{analysis_name}</div>
            <div class="card-sub">Project: {project_name} &nbsp;·&nbsp; Analysis SID: <strong>{analysis_sid}</strong></div>
        </div>""", unsafe_allow_html=True)

        if st.button("🚀  Fetch Results"):
            with st.spinner(f"Fetching results for '{analysis_name}'..."):
                from touchstone_client import get_all_loss_data
                results = get_all_loss_data(analysis_sid)
                st.session_state["last_results"] = results
                total = sum(len(v) for v in results.values() if not v.empty)
                log_pull(project_name, analysis_sid, analysis_name, {k: len(v) for k, v in results.items()})
                if total > 0:
                    st.success(f"✓ {total:,} total records fetched")
                else:
                    st.warning("⚠ No data returned for this analysis")

        if st.session_state.get("last_results"):
            results    = st.session_state["last_results"]
            df_elt     = results.get("ELT",          pd.DataFrame())
            df_ep      = results.get("EP Curves",     pd.DataFrame())
            df_summary = results.get("Loss Summary",  pd.DataFrame())

            st.markdown(f"""
            <div class="metric-row">
                <div class="metric-card blue"><div class="label">ELT records</div><div class="value">{len(df_elt):,}</div><div class="sub">Event loss table</div></div>
                <div class="metric-card green"><div class="label">EP Curves records</div><div class="value">{len(df_ep):,}</div><div class="sub">Annual EP data</div></div>
                <div class="metric-card amber"><div class="label">Loss Summary records</div><div class="value">{len(df_summary):,}</div><div class="sub">EP distribution</div></div>
            </div>""", unsafe_allow_html=True)

            tab1, tab2, tab3 = st.tabs([
                f"📋  ELT ({len(df_elt):,})",
                f"📈  EP Curves ({len(df_ep):,})",
                f"📄  Loss Summary ({len(df_summary):,})"
            ])
            with tab1:
                if not df_elt.empty:
                    st.dataframe(df_elt, use_container_width=True, height=400)
                else:
                    st.info("No ELT data returned for this analysis")
            with tab2:
                if not df_ep.empty:
                    st.dataframe(df_ep, use_container_width=True, height=400)
                else:
                    st.info("No EP Curves data returned for this analysis")
            with tab3:
                if not df_summary.empty:
                    st.dataframe(df_summary, use_container_width=True, height=400)
                else:
                    st.info("No Loss Summary data returned for this analysis")

            st.markdown('<div class="section-header">Download Report</div>', unsafe_allow_html=True)
            if any(not v.empty for v in results.values()):
                excel_bytes = build_excel(analysis_sid, analysis_name, project_name, {
                    "ELT": df_elt, "EP Curves": df_ep, "Loss Summary": df_summary
                })
                safe_name = "".join(c for c in analysis_name if c.isalnum() or c in " _-")[:40].strip()
                filename  = f"Touchstone_{safe_name}_{analysis_sid}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                col1, col2 = st.columns([1, 4])
                with col1:
                    st.download_button("⬇  Download Excel Report", data=excel_bytes, file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                with col2:
                    st.caption(f"📄 {filename}")
            else:
                st.warning("No data available to download")


# ════════════════════════════════════════════════════════════════════════════
# HISTORY
# ════════════════════════════════════════════════════════════════════════════
elif page == "📋  History":
    st.markdown("""
    <div class="page-header">
        <h1>Pull History</h1>
        <p>Record of all results fetched this session</p>
    </div>""", unsafe_allow_html=True)

    history = st.session_state.get("history", [])
    if history:
        st.dataframe(pd.DataFrame(history), use_container_width=True, height=400)
        st.markdown(f"""
        <div class="card">
            <div class="card-title">Session total</div>
            <div class="card-value">{len(history)}</div>
            <div class="card-sub">Analyses pulled this session</div>
        </div>""", unsafe_allow_html=True)
    else:
        st.info("No results pulled yet. Go to Projects to get started.")