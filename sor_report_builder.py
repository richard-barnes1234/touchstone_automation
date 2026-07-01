# sor_report_builder.py
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from model_code_ref import MODEL_CODE_REF
from model_lookup import get_model_description

N_YEARS = 10000


def _thin():
    return Border(*(Side(style="thin", color="D1D5DB"),) * 4)


def _style_header(ws, row, col_start, col_end, fill="1E3A5F"):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin()


def _add_model_code_ref(wb):
    if "ModelCodeRef" in wb.sheetnames:
        return
    ws = wb.create_sheet("ModelCodeRef")
    headers = ["ModelCode", "Model", None, "BriefName", "Analysis Name for Report"]
    for c_idx, h in enumerate(headers, start=1):
        if h:
            ws.cell(row=3, column=c_idx, value=h)
    _style_header(ws, 3, 1, 5)
    for r_idx, (code, model, brief, label) in enumerate(MODEL_CODE_REF, start=4):
        ws.cell(row=r_idx, column=1, value=code)
        ws.cell(row=r_idx, column=2, value=model)
        ws.cell(row=r_idx, column=4, value=brief)
        ws.cell(row=r_idx, column=5, value=label)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 22


def _prepare_df(df_elt):
    """Filter to STC, coerce numeric types. Handles actual API column names."""
    if df_elt is None or df_elt.empty:
        return pd.DataFrame()
    df = df_elt.copy()
    # Normalise column names
    col_map = {}
    for col in df.columns:
        col_lower = col.lower().replace(" ", "").replace("_", "")
        if col_lower == "grossloss" and "sd" not in col_lower:
            col_map[col] = "GrossLoss"
        elif col_lower == "grounduploss" and "sd" not in col_lower:
            col_map[col] = "GroundUpLoss"
        elif col_lower == "eventid":
            col_map[col] = "EventID"
        elif col_lower == "modelcode":
            col_map[col] = "ModelCode"
        elif col_lower == "yearid":
            col_map[col] = "YearID"
        elif col_lower == "catalogtypecode":
            col_map[col] = "CatalogTypeCode"
        elif col_lower == "eventdescription":
            col_map[col] = "EventDescription"
    if col_map:
        df = df.rename(columns=col_map)
    if 'CatalogTypeCode' in df.columns:
        df = df[df['CatalogTypeCode'] == 'STC'].copy()
    for col in ("EventID", "GrossLoss", "GroundUpLoss", "ModelCode", "YearID"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    required = [c for c in ["YearID", "GroundUpLoss", "GrossLoss"] if c in df.columns]
    if required:
        df = df.dropna(subset=required)
    unique_years = df["YearID"].nunique() if not df.empty else 0
    max_year_id  = int(df["YearID"].max()) if not df.empty else N_YEARS
    print(f"  [SOR] Events: {len(df):,} | Unique years: {unique_years:,} | max YearID: {max_year_id:,}")
    if df.empty:
        print("  [SOR] WARNING: No STC events — analysis may use RDS/HIS catalog only")
    return df


def _build_loss_sheet(wb, df, meta):
    """Builds the LossTables sheet. Sheet named after the model label."""
    # N_years per Jennifer's rule:
    # Terrorism (ModelCode 1) = 500,000 years
    # All other perils        = 10,000 years
    # Do NOT use max(YearID) — it can differ from the true simulation length
    n_years = N_YEARS  # default 10,000
    if not df.empty and "ModelCode" in df.columns:
        first_code = df["ModelCode"].dropna().iloc[0] if not df["ModelCode"].dropna().empty else None
        if first_code and int(first_code) == 1:  # ModelCode 1 = AIR Terrorism
            n_years = 500000
    print(f"  [SOR] N_years: {n_years:,}")

    # Model label for sheet name and U1
    model_label = ""
    if not df.empty and "ModelCode" in df.columns:
        first_code = df["ModelCode"].dropna().iloc[0] if not df["ModelCode"].dropna().empty else None
        if first_code:
            from model_code_ref import get_report_label
            model_label = get_report_label(int(first_code)) or get_model_description(int(first_code))

    sheet_name = (model_label or meta.get("analysis_name", "LossTables"))[:31].strip() or "LossTables"
    ws = wb.create_sheet(title=sheet_name)
    last_occ_row = 3 + n_years

    # ── Section headers row 1 ────────────────────────────────────────────────
    ws["I1"] = "AGG"
    ws["O1"] = "OCC"
    ws["I1"].font = Font(bold=True, size=11, color="1E3A5F")
    ws["O1"].font = Font(bold=True, size=11, color="1E3A5F")
    ws["U1"] = "No STC events — analysis may use RDS/HIS catalog only" if df.empty \
               else "=VLOOKUP($F4,ModelCodeRef!$A$3:$E$26,5)"
    ws["U1"].font = Font(bold=True, size=12, color="1E3A5F")

    # ── Column headers row 3 ─────────────────────────────────────────────────
    headers = {
        "A3": "CatalogTypeCode", "B3": "EventDescription", "C3": "EventID",
        "D3": "GrossLoss",       "E3": "GroundUpLoss",     "F3": "ModelCode", "G3": "YearID",
        "I3": "Year", "J3": "GULoss", "K3": "GUStdDevSq", "L3": "GRLoss", "M3": "GRStdDevSq",
        "O3": "Year", "P3": "GURP",   "Q3": "GULoss",     "R3": "GRRP",   "S3": "GRLoss",
        "U3": "Perspective", "V3": 100, "W3": 250, "X3": 500, "Y3": 1000, "Z3": "AAL", "AA3": "SD",
    }
    for coord, val in headers.items():
        ws[coord] = val
    _style_header(ws, 3, 1,  7)
    _style_header(ws, 3, 9,  13)
    _style_header(ws, 3, 15, 19)
    _style_header(ws, 3, 21, 27)

    # ── Raw STC event rows columns A-G ───────────────────────────────────────
    raw_cols   = ["CatalogTypeCode", "EventDescription", "EventID",
                  "GrossLoss", "GroundUpLoss", "ModelCode", "YearID"]
    # Only keep columns that actually exist in the dataframe
    raw_cols   = [c for c in raw_cols if c in df.columns]
    int_cols   = {"EventID", "ModelCode", "YearID"}
    float_cols = {"GrossLoss", "GroundUpLoss"}
    if not df.empty:
        for r_idx, row in enumerate(df[raw_cols].itertuples(index=False), start=4):
            for c_idx, (col_name, val) in enumerate(zip(raw_cols, row), start=1):
                if col_name in int_cols:
                    try: val = int(float(val))
                    except (ValueError, TypeError): pass
                elif col_name in float_cols:
                    try: val = float(val)
                    except (ValueError, TypeError): pass
                ws.cell(row=r_idx, column=c_idx, value=val)

    # ── Pre-calculate ALL values in pandas — no Excel formulas ──────────────
    # This eliminates the "Calculating (threads)..." delay when opening the file.
    # All values written as static numbers — instant open regardless of file size.

    print(f"  [SOR] Pre-calculating AGG/OCC tables in pandas...")

    if not df.empty and "YearID" in df.columns:
        # AGG: sum all losses per year
        agg = df.groupby("YearID").agg(
            GULoss=("GroundUpLoss", "sum"),
            GRLoss=("GrossLoss",    "sum"),
        ).reset_index()
        agg_gu_map = dict(zip(agg["YearID"].astype(int), agg["GULoss"]))
        agg_gr_map = dict(zip(agg["YearID"].astype(int), agg["GRLoss"]))

        # OCC: max single event per year
        occ = df.groupby("YearID").agg(
            GUMax=("GroundUpLoss", "max"),
            GRMax=("GrossLoss",    "max"),
        ).reset_index()
        occ_gu_map = dict(zip(occ["YearID"].astype(int), occ["GUMax"]))
        occ_gr_map = dict(zip(occ["YearID"].astype(int), occ["GRMax"]))

        # AAL: total losses / n_years
        agg_aal_gu = float(df["GroundUpLoss"].sum()) / n_years
        agg_aal_gr = float(df["GrossLoss"].sum())    / n_years
        occ_aal_gu = float(occ["GUMax"].sum())       / n_years
        occ_aal_gr = float(occ["GRMax"].sum())        / n_years

        # SD: standard deviation of annual losses
        all_years_gu = [agg_gu_map.get(y, 0) for y in range(1, n_years + 1)]
        all_years_gr = [agg_gr_map.get(y, 0) for y in range(1, n_years + 1)]
        all_years_occ_gu = [occ_gu_map.get(y, 0) for y in range(1, n_years + 1)]
        all_years_occ_gr = [occ_gr_map.get(y, 0) for y in range(1, n_years + 1)]
        import numpy as np
        agg_sd_gu = float(np.std(all_years_gu, ddof=1))
        agg_sd_gr = float(np.std(all_years_gr, ddof=1))
        occ_sd_gu = float(np.std(all_years_occ_gu, ddof=1))
        occ_sd_gr = float(np.std(all_years_occ_gr, ddof=1))

        # Return period values: rank-based LARGE equivalent
        # Sort descending, pick Nth value (100yr=rank100, 250yr=rank40, etc.)
        rp_map = {100: 100, 250: 40, 500: 20, 1000: 10}

        def rp_val(sorted_desc, rank):
            return round(sorted_desc[rank-1]) if rank <= len(sorted_desc) else 0

        agg_gu_sorted = sorted(agg_gu_map.values(), reverse=True)
        agg_gr_sorted = sorted(agg_gr_map.values(), reverse=True)
        occ_gu_sorted = sorted(occ_gu_map.values(), reverse=True)
        occ_gr_sorted = sorted(occ_gr_map.values(), reverse=True)

        # OCC return period also needs rank-based sort
        occ_gu_sorted_rp = sorted(occ_gu_map.values(), reverse=True)
        occ_gr_sorted_rp = sorted(occ_gr_map.values(), reverse=True)

    else:
        agg_gu_map = agg_gr_map = occ_gu_map = occ_gr_map = {}
        agg_aal_gu = agg_aal_gr = occ_aal_gu = occ_aal_gr = 0
        agg_sd_gu  = agg_sd_gr  = occ_sd_gu  = occ_sd_gr  = 0
        agg_gu_sorted = agg_gr_sorted = []
        occ_gu_sorted_rp = occ_gr_sorted_rp = []
        rp_map = {100: 100, 250: 40, 500: 20, 1000: 10}

    # ── Write AGG table cols I-M (static values) ──────────────────────────────
    for year in range(1, n_years + 1):
        r      = 3 + year
        gu     = agg_gu_map.get(year, 0)
        gr     = agg_gr_map.get(year, 0)
        ws.cell(row=r, column=9,  value=year)
        ws.cell(row=r, column=10, value=round(gu))
        ws.cell(row=r, column=11, value=round((gu - agg_aal_gu) ** 2))
        ws.cell(row=r, column=12, value=round(gr))
        ws.cell(row=r, column=13, value=round((gr - agg_aal_gr) ** 2))

    # ── Write OCC table cols O-S (static values) ──────────────────────────────
    # Sort years by OCC GU descending for return period assignment
    occ_gu_sorted_years = sorted(occ_gu_map.items(), key=lambda x: x[1], reverse=True)
    occ_gu_rank = {year: rank+1 for rank, (year, _) in enumerate(occ_gu_sorted_years)}
    occ_gr_sorted_years = sorted(occ_gr_map.items(), key=lambda x: x[1], reverse=True)
    occ_gr_rank = {year: rank+1 for rank, (year, _) in enumerate(occ_gr_sorted_years)}

    for year in range(1, n_years + 1):
        r      = 3 + year
        gu_max = occ_gu_map.get(year, 0)
        gr_max = occ_gr_map.get(year, 0)
        gu_rp  = round(n_years / occ_gu_rank[year], 1) if year in occ_gu_rank else 0
        gr_rp  = round(n_years / occ_gr_rank[year], 1) if year in occ_gr_rank else 0
        ws.cell(row=r, column=15, value=year)
        ws.cell(row=r, column=16, value=gu_rp)
        ws.cell(row=r, column=17, value=round(gu_max))
        ws.cell(row=r, column=18, value=gr_rp)
        ws.cell(row=r, column=19, value=round(gr_max))

    print(f"  [SOR] AGG/OCC tables written as static values — no recalculation needed")

    # ── EP/AAL/SD summary rows 5-8 — all static values ───────────────────────
    ws["T4"] = "Agg/Occ"
    ws["U4"] = "Perspective"
    ws["T5"] = "AGG"
    ws["U5"] = "Ground Up"
    ws["T6"] = ""
    ws["U6"] = "Gross"
    ws["T7"] = "OCC"
    ws["U7"] = "Ground Up"
    ws["T8"] = ""
    ws["U8"] = "Gross"

    for r in [4, 5, 6, 7, 8]:
        for c in [20, 21]:
            cell = ws.cell(row=r, column=c)
            cell.font      = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="1E3A5F" if r in [4,5,6] else "2E5B9A")
            cell.border    = _thin()
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws["V3"] = 100; ws["W3"] = 250; ws["X3"] = 500; ws["Y3"] = 1000
    ws["Z3"] = "AAL"; ws["AA3"] = "SD"
    _style_header(ws, 3, 22, 27)

    # Static EP values — no LARGE/SUM formulas
    ep_data = {
        5: (agg_gu_sorted, agg_gr_sorted, round(agg_aal_gu), round(agg_sd_gu),
            round(agg_aal_gr), round(agg_sd_gr)),
        6: (agg_gr_sorted, agg_gr_sorted, round(agg_aal_gr), round(agg_sd_gr),
            round(agg_aal_gr), round(agg_sd_gr)),
        7: (occ_gu_sorted_rp, occ_gr_sorted_rp, round(occ_aal_gu), round(occ_sd_gu),
            round(occ_aal_gr), round(occ_sd_gr)),
        8: (occ_gr_sorted_rp, occ_gr_sorted_rp, round(occ_aal_gr), round(occ_sd_gr),
            round(occ_aal_gr), round(occ_sd_gr)),
    }

    for ep_row, rp_col, sorted_list, aal, sd in [
        (5,  "V", agg_gu_sorted,     round(agg_aal_gu), round(agg_sd_gu)),
        (6,  "V", agg_gr_sorted,     round(agg_aal_gr), round(agg_sd_gr)),
        (7,  "V", occ_gu_sorted_rp,  round(occ_aal_gu), round(occ_sd_gu)),
        (8,  "V", occ_gr_sorted_rp,  round(occ_aal_gr), round(occ_sd_gr)),
    ]:
        for rp, rank in [(100, 100), (250, 40), (500, 20), (1000, 10)]:
            col = {100: "V", 250: "W", 500: "X", 1000: "Y"}[rp]
            ws[f"{col}{ep_row}"] = rp_val(sorted_list, rank)
        ws[f"Z{ep_row}"]  = aal
        ws[f"AA{ep_row}"] = sd

    # ── Top 5 events table rows 11-16 ────────────────────────────────────────
    # Pairs GU and Gross by the SAME YEAR (top 5 years ranked by GU loss) —
    # NOT by independent rank, which can mismatch events when GU/Gross sort
    # order differs (confirmed by Jennifer — different deductibles per location
    # can make a year rank high on GU but not on Gross, or vice versa).
    ws["T10"] = "Loss Exceedance"
    ws["U10"] = "Ground Up"
    ws["V10"] = "Gross"
    ws["W10"] = "Max Affected States"
    _style_header(ws, 10, 20, 23, fill="2E5B9A")

    thin = _thin()
    # Find top 5 years by OCC Ground Up loss (column Q), preserving the
    # YearID of each so Gross is looked up for that SAME year — not an
    # independently-ranked Gross value from a potentially different year.
    if not df.empty and "YearID" in df.columns:
        occ_by_year = df.groupby("YearID").agg(
            GUMax=("GroundUpLoss", "max"),
            GRMax=("GrossLoss", "max"),
        ).reset_index().sort_values("GUMax", ascending=False).reset_index(drop=True)
    else:
        occ_by_year = pd.DataFrame(columns=["YearID", "GUMax", "GRMax"])

    top5_levels = [10000, 5000, 10000/3, 2500, 2000]
    for i, level in enumerate(top5_levels):
        r = 11 + i
        ws.cell(row=r, column=20, value=level)
        if i < len(occ_by_year):
            year_id = int(occ_by_year.iloc[i]["YearID"])
            gu_val  = round(occ_by_year.iloc[i]["GUMax"])
            gr_val  = round(occ_by_year.iloc[i]["GRMax"])
            ws.cell(row=r, column=21, value=gu_val)
            ws.cell(row=r, column=22, value=gr_val)
            # Python lookup for EventDescription — no XLOOKUP formula needed
            desc = ""
            if not df.empty and "EventDescription" in df.columns:
                match = df[df["YearID"] == year_id]["EventDescription"]
                if not match.empty:
                    desc = str(match.iloc[0])
            ws.cell(row=r, column=23, value=desc)
        else:
            ws.cell(row=r, column=21, value=0)
            ws.cell(row=r, column=22, value=0)
            ws.cell(row=r, column=23, value="")
        for col in range(20, 24):
            ws.cell(row=r, column=col).border = thin
            ws.cell(row=r, column=col).font   = Font(name="Calibri", size=9)

    # ── Apply whole number format to EP summary and Top 5 cells ─────────────
    for r in range(5, 9):       # EP summary rows 5-8
        for c in range(22, 28): # cols V-AA
            ws.cell(row=r, column=c).number_format = "#,##0"
    for r in range(11, 16):     # Top 5 rows 11-15
        for c in range(20, 23): # cols T-V
            ws.cell(row=r, column=c).number_format = "#,##0"
    # Also format AGG/OCC loss cols J, L, Q, S
    for r in range(4, last_occ_row + 1):
        for c in [10, 12, 17, 19]:  # J, L, Q, S
            ws.cell(row=r, column=c).number_format = "#,##0"
    # Format raw ELT GrossLoss and GroundUpLoss cols D, E
    for r in range(4, 4 + len(df)):
        ws.cell(row=r, column=4).number_format = "#,##0"
        ws.cell(row=r, column=5).number_format = "#,##0"

    # ── Column widths ─────────────────────────────────────────────────────────
    for col, width in {"A":14,"B":50,"C":10,"D":14,"E":14,"F":10,"G":9,
                        "I":8,"J":14,"K":16,"L":14,"M":16,
                        "O":8,"P":12,"Q":14,"R":12,"S":14,
                        "T":10,"U":14,"V":14,"W":14,"X":80,"Y":14,"Z":14,"AA":14}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"
    return ws


def build_sor_report(meta, df_elt):
    """Single-analysis SOR report. Returns BytesIO."""
    df = _prepare_df(df_elt)
    wb = Workbook()
    wb.remove(wb.active)
    _build_loss_sheet(wb, df, meta)
    _add_model_code_ref(wb)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_combined_sor_report(meta, analyses):
    """Multi-analysis combined workbook — one sheet per analysis."""
    wb = Workbook()
    wb.remove(wb.active)
    for analysis in analyses:
        sid   = analysis["sid"]
        atype = analysis["type"]
        sname = analysis.get("sheet_name", f"{atype}-{sid}")
        df_raw = analysis.get("df")
        if atype == "LOSS":
            df = df_raw if isinstance(df_raw, pd.DataFrame) else pd.DataFrame()
            _build_loss_sheet(wb, df, {**meta, "analysis_sid": sid,
                                        "analysis_name": analysis.get("name", sname)})
        elif atype == "HAZ":
            ws_haz = wb.create_sheet(title=sname[:31])
            ws_haz["A1"] = f"Hazard Analysis — SID {sid}"
            ws_haz["A1"].font = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
    _add_model_code_ref(wb)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output