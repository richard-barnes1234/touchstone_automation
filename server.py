# server.py — Flask backend for Touchstone Results Dashboard
# Run: python server.py
# Deploy: Azure App Service (Python 3.12)

from flask import Flask, jsonify, request, send_file, render_template_string
import urllib3
import traceback
import pandas as pd
from model_lookup import get_model_description, get_unique_models
from sor_report_builder import build_sor_report
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = Flask(__name__, static_folder='static', template_folder='templates')

# ── Cache ─────────────────────────────────────────────────────────────────────
_cache = {}


# ── Excel builder ─────────────────────────────────────────────────────────────
def format_sheet(ws):
    thin = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    for cell in ws[1]:
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 28
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        bg = "F0F4F8" if row_idx % 2 == 0 else "FFFFFF"
        for cell in row:
            cell.font = Font(name="Calibri", size=9)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_excel(meta, datasets):
    wb = Workbook()
    # Remove default sheet — we only want data sheets, no summary
    wb.remove(wb.active)

    for sheet_name, df in datasets.items():
        if df is not None and not df.empty:
            ws_data = wb.create_sheet(title=sheet_name[:31])
            for col_idx, col_name in enumerate(df.columns, start=1):
                ws_data.cell(row=1, column=col_idx, value=col_name)
            for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws_data.cell(row=row_idx, column=col_idx, value=value)
            format_sheet(ws_data)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── API Routes ────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    with open("templates/index.html", "r", encoding="utf-8") as f:
        return f.read()


@app.route("/api/projects")
def api_projects():
    try:
        if "projects" not in _cache:
            from get_analysis_sids import get_all_projects
            _cache["projects"] = get_all_projects()
        return jsonify({"ok": True, "data": _cache["projects"], "count": len(_cache["projects"])})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/analyses/loss/<project_sid>")
def api_loss_analyses(project_sid):
    try:
        from get_analysis_sids import get_analyses_for_project
        project_name = request.args.get("project_name", "")
        data = get_analyses_for_project(project_sid, project_name)
        return jsonify({"ok": True, "data": data, "count": len(data)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/analyses/model-batch", methods=["POST"])
def api_model_batch():
    """
    Accepts a list of analysis SIDs, fetches ELT for each (in parallel threads),
    extracts the first ModelCode, and returns a map of {sid: model_name}.
    """
    import concurrent.futures
    try:
        body = request.json or {}
        sids = body.get("sids", [])
        if not sids:
            return jsonify({"ok": True, "data": {}})

        from api_client import send_soap_request
        from soap_templates import get_loss_analysis_event_results
        from config import BUSINESS_UNIT_SID, SQL_INSTANCE_SID
        import xml.etree.ElementTree as ET

        def fetch_model_for_sid(sid):
            try:
                soap = get_loss_analysis_event_results(
                    BUSINESS_UNIT_SID, SQL_INSTANCE_SID, sid)
                r = send_soap_request(soap)
                if r.status_code != 200:
                    return sid, None
                root = ET.fromstring(r.text)
                for elem in root.iter():
                    tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
                    if tag == 'EventLoss':
                        for child in elem:
                            ctag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                            if ctag == 'ModelCode' and child.text:
                                return sid, get_model_description(int(child.text))
                return sid, None
            except Exception:
                return sid, None

        results = {}
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as ex:
            futures = {ex.submit(fetch_model_for_sid, sid): sid for sid in sids}
            for fut in concurrent.futures.as_completed(futures):
                sid, model_name = fut.result()
                if model_name:
                    results[str(sid)] = model_name

        return jsonify({"ok": True, "data": results})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/analyses/haz/<project_sid>")
def api_haz_analyses(project_sid):
    try:
        from get_analysis_sids import get_hazard_analyses_for_project
        project_name = request.args.get("project_name", "")
        data = get_hazard_analyses_for_project(project_sid, project_name)
        return jsonify({"ok": True, "data": data, "count": len(data)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/results/loss/<analysis_sid>")
def api_loss_results(analysis_sid):
    try:
        from touchstone_client import get_all_loss_data
        results = get_all_loss_data(int(analysis_sid))
        data        = {}
        models      = []
        model_label = ""
        for k, df in results.items():
            if not df.empty:
                df = df.replace([float('inf'), float('-inf')], None)
                df = df.where(pd.notnull(df), None)
                rows = []
                for row in df.values.tolist():
                    clean = [None if (v != v or v is None) else v for v in row]
                    rows.append(clean)
                data[k] = {
                    "columns": list(df.columns),
                    "rows":    rows,
                    "count":   len(df)
                }
                # Get model label from first ModelCode in ELT
                if k == 'ELT' and 'ModelCode' in df.columns:
                    models     = get_unique_models(df)
                    first_code = df['ModelCode'].dropna().iloc[0] if not df['ModelCode'].dropna().empty else None
                    model_label = get_model_description(first_code) if first_code else ""
        return jsonify({"ok": True, "data": data, "models": models, "model_label": model_label})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/results/haz/<analysis_sid>")
def api_haz_results(analysis_sid):
    try:
        from touchstone_client import get_hazard_results
        results = get_hazard_results(int(analysis_sid))
        data = {}
        for k, df in results.items():
            if not df.empty:
                df = df.replace([float('inf'), float('-inf')], None)
                df = df.where(pd.notnull(df), None)
                rows = []
                for row in df.values.tolist():
                    clean = [None if (v != v or v is None) else v for v in row]
                    rows.append(clean)
                data[k] = {
                    "columns": list(df.columns),
                    "rows":    rows,
                    "count":   len(df)
                }
        return jsonify({"ok": True, "data": data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/download", methods=["POST"])
def api_download():
    try:
        body          = request.json
        meta          = body.get("meta", {})
        raw_datasets  = body.get("datasets", {})
        analysis_type = meta.get("analysis_type", "LOSS")

        if analysis_type == "LOSS":
            # SOR report needs the FULL unfiltered ELT — re-fetch fresh rather
            # than relying on the column-filtered payload sent from the frontend
            # column selector (which is built for preview/export of other formats).
            from touchstone_client import get_all_loss_data
            analysis_sid = meta.get("analysis_sid")
            results = get_all_loss_data(int(analysis_sid))
            df_elt  = results.get("ELT", pd.DataFrame())
            excel_file = build_sor_report(meta, df_elt)
        else:
            datasets = {}
            for name, payload in raw_datasets.items():
                if payload and payload.get("columns"):
                    df = pd.DataFrame(payload["rows"], columns=payload["columns"])
                    datasets[name] = df
            excel_file = build_excel(meta, datasets)

        # Filename: ProjectName SID: AnalysisSid
        # e.g. "Providence_26 SID: 141190.xlsx"
        # Filename always follows: "ProjectName SID: AnalysisSid.xlsx"
        project  = "".join(c for c in meta.get("project_name", "Report") if c.isalnum() or c in " _-").strip()
        sid      = meta.get("analysis_sid", "")
        filename = f"{project} SID: {sid}.xlsx"

        return send_file(
            excel_file,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500




@app.route("/api/download/combined", methods=["POST"])
def api_download_combined():
    try:
        import traceback

        body          = request.json
        meta          = body.get("meta", {})
        analysis_list = body.get("analyses", [])

        if not analysis_list:
            return jsonify({"ok": False, "error": "No analyses provided"}), 400

        from sor_report_builder import build_combined_sor_report, _prepare_df
        from touchstone_client import get_all_loss_data, get_hazard_results

        used_names = set()

        def unique_name(base):
            base = base[:28].strip()
            name = base
            i = 1
            while name in used_names:
                name = f"{base} {i}"
                i += 1
            used_names.add(name)
            return name

        enriched = []
        for a in analysis_list:
            sid   = int(a["sid"])
            atype = a["type"]
            name  = a.get("name", f"{atype}-{sid}")

            if atype == "LOSS":
                results     = get_all_loss_data(sid)
                df_elt      = results.get("ELT", pd.DataFrame())
                df          = _prepare_df(df_elt)
                model_label = ""
                if not df.empty and "ModelCode" in df.columns:
                    first_code = df["ModelCode"].dropna().iloc[0] \
                        if not df["ModelCode"].dropna().empty else None
                    if first_code:
                        model_label = get_model_description(first_code)
                sheet_name = unique_name(model_label or name or f"LOSS-{sid}")
                enriched.append({"sid": sid, "name": name, "type": atype,
                                  "df": df, "sheet_name": sheet_name})
            elif atype == "HAZ":
                haz_data   = get_hazard_results(sid)
                sheet_name = unique_name(f"HAZ {sid}")
                enriched.append({"sid": sid, "name": name, "type": atype,
                                  "df": haz_data, "sheet_name": sheet_name})

        excel_file = build_combined_sor_report(meta, enriched)

        project  = "".join(c for c in meta.get("project_name", "Combined")
                           if c.isalnum() or c in " _-").strip()
        filename = f"{project} Combined_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            excel_file,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
    #